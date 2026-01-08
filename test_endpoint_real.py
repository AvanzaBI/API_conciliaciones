#!/usr/bin/env python3
"""
Script de prueba real del endpoint /conciliacion-unificada/
Simula una llamada POST con archivos reales y verifica la respuesta.
"""
import requests
import sys
import os
from pathlib import Path
import openpyxl
from io import BytesIO

# Configuración
API_URL = "http://localhost:8000/conciliacion-unificada/"
PDF_PATH = "Formato movimiento diario bancolombia.pdf"
EXCEL_PATH = "Movimiento banco noviembre.xlsx"

def test_endpoint():
    """Ejecuta prueba real del endpoint."""
    print("=" * 60)
    print("PRUEBA REAL DEL ENDPOINT /conciliacion-unificada/")
    print("=" * 60)
    
    # Verificar que los archivos existen
    if not os.path.exists(PDF_PATH):
        print(f"❌ ERROR: No se encuentra el archivo PDF: {PDF_PATH}")
        return False
    
    if not os.path.exists(EXCEL_PATH):
        print(f"❌ ERROR: No se encuentra el archivo Excel: {EXCEL_PATH}")
        return False
    
    print(f"✅ Archivo PDF encontrado: {PDF_PATH}")
    print(f"✅ Archivo Excel encontrado: {EXCEL_PATH}")
    print()
    
    # Preparar archivos para la petición
    try:
        with open(PDF_PATH, 'rb') as pdf_file:
            with open(EXCEL_PATH, 'rb') as excel_file:
                files = {
                    'pdf_file': (os.path.basename(PDF_PATH), pdf_file, 'application/pdf'),
                    'contabilidad_file': (os.path.basename(EXCEL_PATH), excel_file, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                }
                
                print(f"📤 Enviando petición POST a {API_URL}...")
                print()
                
                # Realizar petición
                response = requests.post(API_URL, files=files, timeout=60)
                
                # Verificar código de estado
                print(f"📥 Código de respuesta: {response.status_code}")
                
                if response.status_code != 200:
                    print(f"❌ ERROR: El endpoint respondió con código {response.status_code}")
                    print(f"Respuesta: {response.text[:500]}")
                    return False
                
                print("✅ Endpoint respondió con código 200")
                print()
                
                # Verificar tipo de contenido
                content_type = response.headers.get('Content-Type', '')
                print(f"📄 Content-Type: {content_type}")
                
                if 'spreadsheetml' not in content_type and 'excel' not in content_type.lower():
                    print(f"⚠️  WARNING: Content-Type inesperado: {content_type}")
                else:
                    print("✅ Content-Type correcto (Excel)")
                print()
                
                # Verificar tamaño del archivo
                excel_bytes = response.content
                file_size = len(excel_bytes)
                print(f"📊 Tamaño del archivo Excel: {file_size:,} bytes")
                
                if file_size == 0:
                    print("❌ ERROR: El archivo Excel está vacío")
                    return False
                
                print("✅ Archivo Excel generado correctamente")
                print()
                
                # Verificar estructura del Excel
                print("🔍 Verificando estructura del Excel...")
                try:
                    excel_stream = BytesIO(excel_bytes)
                    workbook = openpyxl.load_workbook(excel_stream)
                    
                    sheet_names = workbook.sheetnames
                    print(f"📋 Hojas encontradas: {len(sheet_names)}")
                    print(f"   Nombres: {', '.join(sheet_names)}")
                    
                    # Verificar hojas esperadas
                    expected_sheets = ['Conciliacion', 'Conceptos', 'Gastos Bancarios']
                    missing_sheets = [s for s in expected_sheets if s not in sheet_names]
                    
                    if missing_sheets:
                        print(f"⚠️  WARNING: Faltan hojas esperadas: {missing_sheets}")
                    else:
                        print("✅ Todas las hojas esperadas están presentes")
                    
                    # Verificar contenido de cada hoja
                    print()
                    print("📊 Contenido de las hojas:")
                    for sheet_name in sheet_names:
                        sheet = workbook[sheet_name]
                        max_row = sheet.max_row
                        max_col = sheet.max_column
                        print(f"   - {sheet_name}: {max_row} filas × {max_col} columnas")
                    
                    workbook.close()
                    
                except Exception as e:
                    print(f"❌ ERROR al leer el Excel: {str(e)}")
                    import traceback
                    traceback.print_exc()
                    return False
                
                print()
                print("=" * 60)
                print("✅ PRUEBA COMPLETADA EXITOSAMENTE")
                print("=" * 60)
                return True
                
    except requests.exceptions.ConnectionError:
        print(f"❌ ERROR: No se pudo conectar al servidor en {API_URL}")
        print("   Asegúrate de que el servidor FastAPI esté corriendo:")
        print("   uvicorn main:app --host 0.0.0.0 --port 8000")
        return False
    except requests.exceptions.Timeout:
        print(f"❌ ERROR: La petición excedió el tiempo límite (60s)")
        return False
    except Exception as e:
        print(f"❌ ERROR inesperado: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    success = test_endpoint()
    sys.exit(0 if success else 1)

