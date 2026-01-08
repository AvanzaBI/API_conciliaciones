import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font 
from openpyxl.utils import get_column_letter
from io import BytesIO
from openpyxl.styles import numbers
import re

def _norm(s: str) -> str:
    s = (s or "").strip().upper()
    s = (
        s.replace("Á","A").replace("É","E").replace("Í","I")
         .replace("Ó","O").replace("Ú","U").replace("Ü","U")
    )
    s = re.sub(r"\s+", " ", s)
    return s

def _safe_drop_columns(df: pd.DataFrame, columns: list[str]) -> pd.DataFrame:
    """Elimina columnas solo si existen, evitando errores por columnas faltantes."""
    cols_to_drop = [col for col in columns if col in df.columns]
    return df.drop(cols_to_drop, axis=1) if cols_to_drop else df

def conciliar_movimientos(df_contabilidad: pd.DataFrame, df_extracto: pd.DataFrame) -> bytes:

    df1 = df_contabilidad.copy()
    df2 = df_extracto.copy()

    # mapa normalizado, para soportar "Asiento " o "ASIENTO"
    colmap = { _norm(c): c for c in df1.columns }

    # detectar por nombre real
    fecha_col = colmap.get("FECHA") or colmap.get("Fecha".upper())
    mov_col   = colmap.get("MOVIMIENTO") or colmap.get("Movimiento".upper())
    asiento_col = colmap.get("ASIENTO")

    # Si es el archivo Movimiento banco noviembre.xlsx
    if fecha_col and mov_col:
        df1 = df1.rename(columns={
            fecha_col: "FECHA",
            mov_col: "VALOR",
            **({asiento_col: "Concepto Contabilidad"} if asiento_col else {})
        })
    else:
        # Si ya viene en el formato esperado, valida que existan 3 columnas
        if df1.shape[1] < 3:
            raise ValueError(f"Contabilidad tiene pocas columnas: {list(df1.columns)}")
        # si ya trae FECHA, VALOR, Concepto Contabilidad, no toques
        # si no, aquí es donde deberías mapear tu otro formato, no por posición
        # por ahora, intenta detectar por nombres
        colmap = { _norm(c): c for c in df1.columns }
        fecha_col = colmap.get("FECHA")
        valor_col = colmap.get("VALOR")
        concepto_col = colmap.get("CONCEPTO CONTABILIDAD") or colmap.get("CONCEPTO") or colmap.get("DESCRIPCION")

        if fecha_col and valor_col:
            df1 = df1.rename(columns={
                fecha_col: "FECHA",
                valor_col: "VALOR",
                **({concepto_col: "Concepto Contabilidad"} if concepto_col else {})
            })
        else:
            raise ValueError(f"No detecté FECHA y VALOR en contabilidad. Columnas: {list(df1.columns)}")

    # asegurar columna concepto
    if "Concepto Contabilidad" not in df1.columns:
        df1["Concepto Contabilidad"] = ""

    # quedarnos con las 3 columnas finales
    df1 = df1[["FECHA", "Concepto Contabilidad", "VALOR"]]
    # FIX CRÍTICO: Resetear índice inmediatamente después de seleccionar columnas
    df1 = df1.reset_index(drop=True)

   # --- Preparación del df1 ---
    df1['FECHA'] = pd.to_datetime(df1['FECHA'], dayfirst=True, errors="coerce").dt.strftime("%d/%m/%Y")
    df1['VALOR'] = pd.to_numeric(df1['VALOR'], errors="coerce").fillna(0).astype(int)
    df1['clave_unica'] = df1['FECHA'] + '_' + df1['VALOR'].astype(str)

    # --- Preparación del df2 ---
    # FIX: Resetear índice antes de operaciones para evitar problemas de alineación
    df2 = df2.reset_index(drop=True)
    df2['FECHA'] = pd.to_datetime(df2['FECHA'], dayfirst=True, errors="coerce").dt.strftime("%d/%m/%Y")
    df2['VALOR'] = pd.to_numeric(df2['VALOR'], errors="coerce").fillna(0).astype(int)
    df2['clave_unica'] = df2['FECHA'] + '_' + df2['VALOR'].astype(str)



    merged_df = pd.merge(df1, df2, on='clave_unica', how='outer', suffixes=('_Contabilidad', '_Extracto'))
    # FIX CRÍTICO: Resetear índice para evitar "Unalignable boolean Series"
    merged_df = merged_df.reset_index(drop=True)
    
    # FIX CRÍTICO: Asegurar tipos numéricos para comparaciones seguras
    merged_df['VALOR_Contabilidad'] = pd.to_numeric(merged_df['VALOR_Contabilidad'], errors="coerce")
    merged_df['VALOR_Extracto'] = pd.to_numeric(merged_df['VALOR_Extracto'], errors="coerce")
    
    consolidado = merged_df.copy()

    # FIX CRÍTICO: Usar .values para evitar problemas de alineación de índices
    # Crear máscaras booleanas usando .values para asegurar alineación perfecta
    mask_valor_cont_pos = (merged_df['VALOR_Contabilidad'].values > 0)
    mask_valor_cont_neg = (merged_df['VALOR_Contabilidad'].values < 0)
    mask_valor_ext_pos = (merged_df['VALOR_Extracto'].values > 0)
    mask_valor_ext_neg = (merged_df['VALOR_Extracto'].values < 0)
    mask_valor_cont_na = pd.isna(merged_df['VALOR_Contabilidad'].values)
    mask_valor_ext_na = pd.isna(merged_df['VALOR_Extracto'].values)
    
    # Caso 1: Entradas en contabilidad y no en extracto
    mask_caso1 = mask_valor_cont_pos & mask_valor_ext_na
    caso_1 = merged_df[mask_caso1].copy()
    caso_1 = _safe_drop_columns(caso_1, ['FECHA_Extracto', 'VALOR_Extracto', 'clave_unica', 'DESCRIPCION_Extracto'])
    total_caso1 = caso_1['VALOR_Contabilidad'].sum() if not caso_1.empty else 0
    
    # Caso 2: Entradas en extracto y no en contabilidad
    mask_caso2 = mask_valor_ext_pos & mask_valor_cont_na
    caso_2 = merged_df[mask_caso2].copy()
    caso_2 = _safe_drop_columns(caso_2, ['FECHA_Contabilidad', 'VALOR_Contabilidad', 'Concepto Contabilidad_Contabilidad', 'clave_unica'])
    total_caso2 = caso_2['VALOR_Extracto'].sum() if not caso_2.empty else 0
    
    # Caso 3: Salidas en contabilidad y no en extracto
    mask_caso3 = mask_valor_cont_neg & mask_valor_ext_na
    caso_3 = merged_df[mask_caso3].copy()
    caso_3 = _safe_drop_columns(caso_3, ['FECHA_Extracto', 'VALOR_Extracto', 'DESCRIPCION_Extracto', 'clave_unica'])
    total_caso3 = caso_3['VALOR_Contabilidad'].sum() if not caso_3.empty else 0
    
    # Caso 4: Salidas en extracto y no en contabilidad
    mask_caso4 = mask_valor_ext_neg & mask_valor_cont_na
    caso_4 = merged_df[mask_caso4].copy()
    caso_4 = _safe_drop_columns(caso_4, ['FECHA_Contabilidad', 'VALOR_Contabilidad', 'Concepto Contabilidad_Contabilidad', 'clave_unica'])
    total_caso4 = caso_4['VALOR_Extracto'].sum() if not caso_4.empty else 0
    # Eliminar la columna de clave temporal (solo si existe)
    if 'clave_unica' in consolidado.columns:
        consolidado.drop('clave_unica', axis=1, inplace=True)
    
    # FIX CRÍTICO: Resetear índice antes de seleccionar columnas para evitar problemas de alineación
    consolidado = consolidado.reset_index(drop=True)
    
    # FIX CRÍTICO: Verificar que las columnas existan antes de seleccionarlas
    cols_consolidado = ['FECHA_Contabilidad', 'VALOR_Contabilidad', 'FECHA_Extracto', 'VALOR_Extracto']
    cols_existentes = [col for col in cols_consolidado if col in consolidado.columns]
    if cols_existentes:
        consolidado = consolidado[cols_existentes]
    else:
        # Si no existen las columnas esperadas, crear DataFrame vacío con esas columnas
        consolidado = pd.DataFrame(columns=cols_consolidado)
    # Hoja gastos bancarios
    # FIX: Proteger uso de .isin() verificando que la columna existe y resetear índice
    # Asegurar que df2 tenga índice continuo antes de usar .isin()
    df2 = df2.reset_index(drop=True)
    
    ingresos =["ABONO INTERESES AHORROS","AJUSTE INTERES AHORROS DB"]
    if 'DESCRIPCION' in df2.columns:
        # Usar .values para evitar problemas de alineación
        mask_ingresos = df2['DESCRIPCION'].isin(ingresos).values
        df_ingresos = df2[mask_ingresos].copy()
        if not df_ingresos.empty:
            df_ingresos['VALOR'] = df_ingresos['VALOR'].abs()
            df_ingresos = df_ingresos.groupby('DESCRIPCION').agg({'VALOR':'sum'}).reset_index()
        else:
            df_ingresos = pd.DataFrame(columns=['DESCRIPCION', 'VALOR'])
    else:
        df_ingresos = pd.DataFrame(columns=['DESCRIPCION', 'VALOR'])

    gastos_bancarios=["IMPTO GOBIERNO 4X1000","CUOTA MANEJO SUC VIRT EMPRESA","COMISION PAGO A PROVEEDORES","COMISION PAGO A NOMINA"]
    if 'DESCRIPCION' in df2.columns:
        # Usar .values para evitar problemas de alineación
        mask_gastos = df2['DESCRIPCION'].isin(gastos_bancarios).values
        df_gastos_bancarios = df2[mask_gastos].copy()
        if not df_gastos_bancarios.empty:
            df_gastos_bancarios['VALOR'] = df_gastos_bancarios['VALOR'].abs()
            df_gastos_bancarios = df_gastos_bancarios.groupby('DESCRIPCION').agg({'VALOR':'sum'}).reset_index()
        else:
            df_gastos_bancarios = pd.DataFrame(columns=['DESCRIPCION', 'VALOR'])
    else:
        df_gastos_bancarios = pd.DataFrame(columns=['DESCRIPCION', 'VALOR'])

    impuestos =["IVA CUOTA MANEJO SUC VIRT EMP","COBRO IVA PAGOS AUTOMATICOS"]
    if 'DESCRIPCION' in df2.columns:
        # Usar .values para evitar problemas de alineación
        mask_impuestos = df2['DESCRIPCION'].isin(impuestos).values
        df_impuestos = df2[mask_impuestos].copy()
        if not df_impuestos.empty:
            df_impuestos['VALOR'] = df_impuestos['VALOR'].abs()
            df_impuestos = df_impuestos.groupby('DESCRIPCION').agg({'VALOR':'sum'}).reset_index()
        else:
            df_impuestos = pd.DataFrame(columns=['DESCRIPCION', 'VALOR'])
    else:
        df_impuestos = pd.DataFrame(columns=['DESCRIPCION', 'VALOR'])
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:

            # 🔹 Hoja 1: Resultado del join con formato
            # Escribimos el título primero
            if 'clave_unica' in merged_df.columns:
                merged_df.drop(columns=['clave_unica'], inplace=True)
            merged_df.to_excel(writer, sheet_name='Conciliacion', index=False, startrow=2)
            worksheet = writer.sheets['Conciliacion']
            worksheet.cell(row=1, column=1, value="Resultado de la Conciliación Bancaria").font = Font(bold=True, size=14)

            # Ajustar el ancho de las columnas y formatear fechas (usando openpyxl)
            for col_idx, col in enumerate(worksheet.columns, 1):
                max_length = 0
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

            

            # 🔹 Hoja 2: Ejemplo de encabezados y tablas
            # Creamos un DataFrame de ejemplo
            caso_1.to_excel(writer, sheet_name='Conceptos', index=False, startrow=4)
            caso_2.to_excel(writer, sheet_name='Conceptos', index=False, startrow=4+len(caso_1)+3)
            caso_3.to_excel(writer, sheet_name='Conceptos', index=False, startrow=4+len(caso_1) + len(caso_2) + 6)
            caso_4.to_excel(writer, sheet_name='Conceptos', index=False, startrow=4+len(caso_1) + len(caso_2) + len(caso_3) + 9)

            # Accedemos a la hoja 'Ejemplo' para añadir los textos y encabezados
            worksheet_ejemplo = writer.sheets['Conceptos']
            worksheet_ejemplo.cell(row=1, column=1, value="Formato de Conciliación Bancaria").font = Font(bold=True, size=14)
            worksheet_ejemplo.cell(row=3, column=1, value="Caso 1: Entradas en Contabilidad no en Extracto").font = Font(bold=True)
            worksheet_ejemplo.cell(row=4+len(caso_1)+3, column=1, value="Caso 2: Entradas en Extracto y no en Contabilidad").font = Font(bold=True)
            worksheet_ejemplo.cell(row=4+len(caso_1) + len(caso_2) + 6, column=1, value="Caso 3: Salidas en Contabilidad no en Extracto").font = Font(bold=True)
            worksheet_ejemplo.cell(row=4+len(caso_1) + len(caso_2) + len(caso_3) + 9, column=1, value="Caso 4: Salidas en Extracto y no en Contabilidad").font = Font(bold=True)
            worksheet_ejemplo.cell(row=5, column=4, value=f"Total Caso 1: ${total_caso1}").font = Font(bold=True)
            worksheet_ejemplo.cell(row=5+len(caso_1)+2, column=4, value=f"Total Caso 2: ${total_caso2}").font = Font(bold=True)
            worksheet_ejemplo.cell(row=5+len(caso_1) + len(caso_2) + 5, column=4, value=f"Total Caso 3: ${total_caso3}").font = Font(bold=True)
            worksheet_ejemplo.cell(row=5+len(caso_1) + len(caso_2) + len(caso_3) + 7, column=4, value=f"Total Caso 4: ${total_caso4}").font = Font(bold=True)

            # Ajustar el ancho de las columnas y formatear fechas (usando openpyxl)
            for col_idx, col in enumerate(worksheet_ejemplo.columns, 1):
                max_length = 0
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                worksheet_ejemplo.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

            # Buscar encabezados y aplicar formato en columnas de valores
            for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row):
                for cell in row:
                    if cell.value and "VALOR" in str(cell.value).upper():
                        col_idx = cell.col_idx  # número de columna de este encabezado
            
                        # Aplicar formato desde la fila siguiente hasta el final
                        for data_row in range(cell.row + 1, worksheet.max_row + 1):
                            valor_cell = worksheet.cell(data_row, col_idx)
                            if isinstance(valor_cell.value, (int, float)):
                                valor_cell.number_format = '"$"#,##0'  # COP sin decimales


            # Formatear columnas de VALOR como moneda en todas las hojas# Recorremos todas las filas
            for row_idx, row in enumerate(worksheet_ejemplo.iter_rows(min_row=1, max_row=worksheet_ejemplo.max_row), start=1):
                first_cell = str(row[0].value).strip() if row[0].value else ""

                # Detectamos inicio de tabla
                if first_cell.startswith("Caso"):
                    # La fila 2 después del encabezado de caso suele ser encabezados de tabla
                    header_row = row_idx + 1

                    # Recorremos desde header_row+1 hasta encontrar fila vacía
                    data_row = header_row + 1
                    while data_row <= worksheet_ejemplo.max_row and worksheet_ejemplo.cell(data_row, 1).value:
                        cell = worksheet_ejemplo.cell(data_row, 3)  # Columna C (VALOR)
                        if isinstance(cell.value, (int, float)):
                            cell.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
                        data_row += 1

            # Hoja 3: Gastos Bancarios
            df_ingresos.to_excel(writer, sheet_name='Gastos Bancarios', index=False, startrow=2)
            df_gastos_bancarios.to_excel(writer, sheet_name='Gastos Bancarios', index=False, startrow=2+len(df_ingresos)+4)
            df_impuestos.to_excel(writer, sheet_name='Gastos Bancarios', index=False, startrow=2+len(df_ingresos)+4+len(df_gastos_bancarios)+4)

            # Textos y encabezados
            worksheet_gastos = writer.sheets['Gastos Bancarios']
            worksheet_gastos.cell(row=1, column=1, value="Gastos Bancarios").font = Font(bold=True, size=14)
            worksheet_gastos.cell(row=2, column=1, value="Ingresos").font = Font(bold=True)
            worksheet_gastos.cell(row=2+len(df_ingresos)+4, column=1, value="Gastos Bancarios").font = Font(bold=True)
            worksheet_gastos.cell(row=2+len(df_ingresos)+4+len(df_gastos_bancarios)+4, column=1, value="Impuestos").font = Font(bold=True)

            # Ajustar el ancho de las columnas y formatear fechas (usando openpyxl)
            for col_idx, col in enumerate(worksheet_gastos.columns, 1):
                max_length = 0
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                worksheet_gastos.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

            # Recorremos filas para encontrar encabezados de tabla
            for row_idx, row in enumerate(worksheet_gastos.iter_rows(min_row=1, max_row=worksheet_gastos.max_row), start=1):
                # Detectar encabezados de tabla (cuando la segunda celda dice "VALOR")
                if row[1].value and str(row[1].value).strip().upper() == "VALOR":
                    # Empezar a leer los datos desde la siguiente fila
                    data_row = row_idx + 1

                    # Recorremos hasta encontrar una fila vacía (fin de la tabla)
                    while data_row <= worksheet_gastos.max_row and worksheet_gastos.cell(data_row, 1).value:
                        cell = worksheet_gastos.cell(data_row, 2)  # Columna B (VALOR)
                        if isinstance(cell.value, (int, float)):
                            # Formato pesos colombianos sin decimales
                            cell.number_format = '"$"#,##0'
                        data_row += 1


    # Guardar el archivo Excel en memoria
    output.seek(0)
    return output.read()  # Retornamos los bytes del archivo Excel
